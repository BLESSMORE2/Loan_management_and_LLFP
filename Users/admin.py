from django.contrib import admin
from django.contrib.auth.admin import UserAdmin
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter
from .models import CustomUser, AuditTrail
class CustomUserAdmin(UserAdmin):
    # The fields to be used in displaying the User model in the admin interface
    fieldsets = (
        (None, {'fields': ('email', 'password')}),
        ('Personal info', {'fields': ('name', 'surname', 'phone_number', 'address', 'department', 'gender')}),
        ('Permissions', {'fields': ('is_active', 'is_staff', 'is_superuser', 'groups', 'user_permissions')}),
        ('Important dates', {'fields': ('last_login', 'date_joined')}),
    )

    # Add fields for adding/editing in the admin interface
    add_fieldsets = (
        (None, {
            'classes': ('wide',),
            'fields': ('email', 'surname', 'password1', 'password2', 'is_active', 'is_staff', 'is_superuser')}
        ),
    )

    # Define the list of displayed fields in the user listing
    list_display = ('email', 'name', 'surname', 'is_staff', 'is_active')
    search_fields = ('email', 'name', 'surname')
    ordering = ('email',)

# Register the custom admin class with the CustomUser model
admin.site.register(CustomUser, CustomUserAdmin)


@admin.register(AuditTrail)
class AuditTrailAdmin(admin.ModelAdmin):
    list_display = ('user__email', 'model_name', 'action', 'object_id', 'timestamp')
    list_filter = ('model_name', 'action', 'timestamp')
    search_fields = ('user__email', 'model_name', 'object_id', 'change_description')
    ordering = ('-timestamp',)

    # Add custom actions
    actions = ['export_as_excel']

    def export_as_excel(self, request, queryset):
        """
        Export selected audit trail records to an Excel file.
        """
        # Create a workbook and sheet
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Audit Trails"

        # Define column headers
        headers = ['User Email', 'Model Name', 'Action', 'Object ID', 'Timestamp', 'Change Description']
        sheet.append(headers)

        # Add data rows
        for record in queryset:
            row = [
                record.user.email if record.user else 'N/A',
                record.model_name,
                record.action,
                record.object_id,
                record.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                record.change_description,
            ]
            sheet.append(row)

        # Adjust column widths
        for col_num, _ in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            sheet.column_dimensions[col_letter].width = 25

        # Create HTTP response with Excel file
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="AuditTrails.xlsx"'
        workbook.save(response)

        return response

    export_as_excel.short_description = "Export selected records to Excel"