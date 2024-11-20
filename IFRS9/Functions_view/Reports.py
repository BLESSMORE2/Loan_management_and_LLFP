from ast import Import
from django.shortcuts import render,redirect, get_object_or_404
from ..models import FCT_Reporting_Lines, ReportColumnConfig
from django.db.models import Max
from django.contrib import messages
import csv
from django.http import HttpResponse
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.db.models import Q
from django.contrib.auth.decorators import login_required
import pandas as pd
import numpy as np
from datetime import datetime
import os
from openpyxl import Workbook
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Max, Sum, Count, Q


@login_required
def reporting_home(request):
    return render(request, 'reports/reporting.html')

@login_required
def list_reports(request):
    # This view will render the list of available reports
    return render(request, 'reports/list_reports.html')




@login_required
@require_http_methods(["GET", "POST"])
def view_results_and_extract(request):
    # Handle AJAX request for dynamic Run Key loading (unchanged)
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest' and request.GET.get('field') == 'fic_mis_date':
        fic_mis_date = request.GET.get('value')
        n_run_keys = FCT_Reporting_Lines.objects.filter(fic_mis_date=fic_mis_date).order_by('-n_run_key').values_list('n_run_key', flat=True).distinct()
        return JsonResponse({'n_run_keys': list(n_run_keys) if n_run_keys.exists() else []})

    # Non-AJAX processing
    fic_mis_dates = FCT_Reporting_Lines.objects.order_by('-fic_mis_date').values_list('fic_mis_date', flat=True).distinct()
    fic_mis_date = request.GET.get('fic_mis_date')
    n_run_key = request.GET.get('n_run_key')

    # Save selected filters to the session
    if fic_mis_date and n_run_key:
        request.session['fic_mis_date'] = fic_mis_date
        request.session['n_run_key'] = n_run_key
    else:
        request.session.pop('fic_mis_date', None)
        request.session.pop('n_run_key', None)

    # Ensure filters are provided
    if not fic_mis_date or not n_run_key:
        messages.error(request, "Both Reporting Date and Run Key are required.")
        return render(request, 'reports/report_view.html', {
            'selected_columns': [],
            'report_data': [],
            'filters': request.GET,
            'fic_mis_dates': fic_mis_dates,
            'fic_mis_date': fic_mis_date,
            'n_run_key': n_run_key,
        })

    # Apply filters and fetch data (unchanged)
    filters = {'fic_mis_date': fic_mis_date, 'n_run_key': n_run_key}
    report_config = get_object_or_404(ReportColumnConfig, report_name="default_report")
    selected_columns = report_config.selected_columns
    report_data = FCT_Reporting_Lines.objects.filter(**filters).values(*selected_columns)

    paginator = Paginator(report_data, 25)
    page = request.GET.get('page', 1)
    try:
        paginated_report_data = paginator.page(page)
    except (PageNotAnInteger, EmptyPage):
        paginated_report_data = paginator.page(1)

    return render(request, 'reports/report_view.html', {
        'selected_columns': selected_columns,
        'report_data': paginated_report_data,
        'filters': filters,
        'fic_mis_dates': fic_mis_dates,
        'fic_mis_date': fic_mis_date,
        'n_run_key': n_run_key,
    })


@login_required
def download_report(request):
    # Retrieve filters from the session
    fic_mis_date = request.session.get('fic_mis_date')
    n_run_key = request.session.get('n_run_key')

    # Validate filters
    if not fic_mis_date or not n_run_key:
        return HttpResponse("Both Reporting Date and Run Key are required.", status=400)

    # Apply filters to fetch data
    filters = {'fic_mis_date': fic_mis_date, 'n_run_key': n_run_key}
    try:
        report_config = ReportColumnConfig.objects.get(report_name="default_report")
        selected_columns = report_config.selected_columns
    except ReportColumnConfig.DoesNotExist:
        return HttpResponse("Report configuration not found.", status=404)

    report_data = FCT_Reporting_Lines.objects.filter(**filters).values(*selected_columns)
    if not report_data.exists():
        return HttpResponse("No data found for the selected filters.", status=404)

    # Create CSV response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="filtered_report.csv"'
    writer = csv.writer(response)
    writer.writerow(selected_columns)

    for row in report_data:
        writer.writerow([row[column] for column in selected_columns])

    return response


############################################

# Directory to save the CSV files
CSV_DIR = os.path.join(os.getcwd(), 'csv_files')

@login_required
@require_http_methods(["GET", "POST"])
def ecl_main_filter_view(request):
    # Handle POST request (applying the filter)
    if request.method == 'POST':
        # Retrieve selected Reporting Date and Run Key from the form
        fic_mis_date = request.POST.get('fic_mis_date')
        n_run_key = request.POST.get('n_run_key')
        
        if fic_mis_date and n_run_key:
            # Store the selected filter values in session for later use
            request.session['fic_mis_date'] = fic_mis_date
            request.session['n_run_key'] = n_run_key

            # Retrieve filtered data based on the selected main filter values
            ecl_data = FCT_Reporting_Lines.objects.filter(fic_mis_date=fic_mis_date, n_run_key=n_run_key)

            # Convert the filtered data into a DataFrame
            ecl_data_df = pd.DataFrame(list(ecl_data.values()))

            # Convert date fields to strings
            if 'fic_mis_date' in ecl_data_df.columns:
                ecl_data_df['fic_mis_date'] = ecl_data_df['fic_mis_date'].astype(str)
            if 'd_maturity_date' in ecl_data_df.columns:
                ecl_data_df['d_maturity_date'] = ecl_data_df['d_maturity_date'].astype(str)

            # Create the directory if it doesn't exist
            if not os.path.exists(CSV_DIR):
                os.makedirs(CSV_DIR)

            # Save the data as a CSV file in the session (store the filename)
            csv_filename = os.path.join(CSV_DIR, f"ecl_data_{fic_mis_date}_{n_run_key}.csv")
            ecl_data_df.to_csv(csv_filename, index=False)
            request.session['csv_filename'] = csv_filename

            # Redirect to the sub-filter view
            return redirect('ecl_sub_filter_view')

    # Handle GET request
    fic_mis_dates = FCT_Reporting_Lines.objects.order_by('-fic_mis_date').values_list('fic_mis_date', flat=True).distinct()

    # Check for AJAX requests to dynamically update run key dropdown
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        field_name = request.GET.get('field')
        field_value = request.GET.get('value')

        if field_name == 'fic_mis_date':
            # Fetch run keys corresponding to the selected FIC MIS date
            n_run_keys = FCT_Reporting_Lines.objects.filter(fic_mis_date=field_value).order_by('-n_run_key').values_list('n_run_key', flat=True).distinct()
            return JsonResponse({'n_run_keys': list(n_run_keys)})

    # Render the main filter template with Reporting Dates
    return render(request, 'reports/ecl_summary_report.html', {'fic_mis_dates': fic_mis_dates})




@login_required
@require_http_methods(["GET", "POST"])
def ecl_sub_filter_view(request):
    # Retrieve the CSV filename from the session
    csv_filename = request.session.get('csv_filename')

    # If no data is available, redirect to the main filter page
    if not csv_filename or not os.path.exists(csv_filename):
        return redirect('ecl_main_filter_view')

    # Load the data from the CSV file
    ecl_data_df = pd.read_csv(csv_filename)

    # Retrieve sub-filter form fields from the request
    n_prod_segment = request.GET.get('n_prod_segment')
    n_prod_type = request.GET.get('n_prod_type')
    n_stage_descr = request.GET.get('n_stage_descr')
    n_loan_type = request.GET.get('n_loan_type')

    # Apply sub-filters if provided
    if n_prod_segment:
        ecl_data_df = ecl_data_df[ecl_data_df['n_prod_segment'] == n_prod_segment]
    if n_prod_type:
        ecl_data_df = ecl_data_df[ecl_data_df['n_prod_type'] == n_prod_type]
    if n_stage_descr:
        ecl_data_df = ecl_data_df[ecl_data_df['n_stage_descr'] == n_stage_descr]
    if n_loan_type:
        ecl_data_df = ecl_data_df[ecl_data_df['n_loan_type'] == n_loan_type]

    # Retrieve the selected group by field from the request (GET or POST)
    group_by_field = request.GET.get('group_by_field', 'n_stage_descr')  # Default group by 'n_stage_descr'

    # Group the data by the selected field and sum the amounts, while also counting unique accounts
    grouped_data = (
        ecl_data_df.groupby(group_by_field)
        .agg({
            'n_exposure_at_default_ncy': 'sum',
            'n_exposure_at_default_rcy': 'sum',
            'n_12m_ecl_rcy': 'sum',
            'n_lifetime_ecl_rcy': 'sum',
            'n_account_number': pd.Series.nunique,  # Count unique accounts in each group
        })
        .reset_index()
        .to_dict(orient='records')
    )

    # Calculate grand totals
    grand_totals = {
        'n_exposure_at_default_ncy': ecl_data_df['n_exposure_at_default_ncy'].sum(),
        'n_exposure_at_default_rcy': ecl_data_df['n_exposure_at_default_rcy'].sum(),
        'n_12m_ecl_rcy': ecl_data_df['n_12m_ecl_rcy'].sum(),
        'n_lifetime_ecl_rcy': ecl_data_df['n_lifetime_ecl_rcy'].sum(),
        'n_account_number': ecl_data_df['n_account_number'].nunique(),  # Grand total for unique accounts
    }

    # Calculate percentages for the second table
    grouped_data_percentages = []
    total_ecl_12m = grand_totals['n_12m_ecl_rcy']
    total_ecl_lifetime = grand_totals['n_lifetime_ecl_rcy']
    total_accounts = grand_totals['n_account_number']

    for row in grouped_data:
        grouped_data_percentages.append({
            group_by_field: row[group_by_field],
            'percent_12m_ecl_rcy': (row['n_12m_ecl_rcy'] / total_ecl_12m * 100) if total_ecl_12m else 0,
            'percent_lifetime_ecl_rcy': (row['n_lifetime_ecl_rcy'] / total_ecl_lifetime * 100) if total_ecl_lifetime else 0,
            'percent_accounts': (row['n_account_number'] / total_accounts * 100) if total_accounts else 0,
        })

    # Distinct values for sub-filters
    distinct_prod_segments = list(ecl_data_df['n_prod_segment'].unique())
    distinct_prod_types = list(ecl_data_df['n_prod_type'].unique())
    distinct_stage_descrs = list(ecl_data_df['n_stage_descr'].unique())
    distinct_loan_types = list(ecl_data_df['n_loan_type'].unique())

    # Store the grouped data and grand totals in the session for Excel export
    request.session['grouped_data'] = grouped_data
    request.session['group_by_field'] = group_by_field
    request.session['grand_totals'] = grand_totals
    request.session['grouped_data_percentages'] = grouped_data_percentages

    # Render the sub-filter view template
    return render(request, 'reports/ecl_summary_report_sub.html', {
        'grouped_data': grouped_data,
        'grouped_data_percentages': grouped_data_percentages,
        'group_by_field': group_by_field,
        'distinct_prod_segments': distinct_prod_segments,
        'distinct_prod_types': distinct_prod_types,
        'distinct_stage_descrs': distinct_stage_descrs,
        'distinct_loan_types': distinct_loan_types,
        'grand_totals': grand_totals,
    })


# Export to Excel dynamically based on the current filtered and grouped data
@login_required
def export_ecl_report_to_excel(request):
    # Retrieve the filtered data, grouped data, grouped percentages, and grand totals from session
    grouped_data = request.session.get('grouped_data', [])
    grouped_data_percentages = request.session.get('grouped_data_percentages', [])
    group_by_field = request.session.get('group_by_field', 'n_stage_descr')
    grand_totals = request.session.get('grand_totals', {})

    # Convert the grouped data into pandas DataFrames
    df_grouped = pd.DataFrame(grouped_data)
    df_percentages = pd.DataFrame(grouped_data_percentages)

    # Check if percentages data exists and add missing fields
    if not df_percentages.empty and group_by_field not in df_percentages.columns:
        df_percentages[group_by_field] = [row[group_by_field] for row in grouped_data]

    # Create a new Excel workbook
    wb = openpyxl.Workbook()

    # Add the first sheet for the absolute numbers
    ws1 = wb.active
    ws1.title = "ECL Report (Absolute)"
    headers = [group_by_field, "EAD Orig Currency", "EAD Reporting Currency", "12 Month Reporting ECL", "Lifetime Reporting ECL", "Number of Accounts"]
    ws1.append(headers)

    # Add the grouped data to the first sheet
    for row in dataframe_to_rows(df_grouped, index=False, header=False):
        ws1.append(row)

    # Add the grand totals at the bottom of the first sheet
    ws1.append([
        'Grand Total',
        grand_totals['n_exposure_at_default_ncy'],
        grand_totals['n_exposure_at_default_rcy'],
        grand_totals['n_12m_ecl_rcy'],
        grand_totals['n_lifetime_ecl_rcy'],
        grand_totals['n_account_number']
    ])

    # Add styling to the first sheet
    style_excel_sheet(ws1, len(grouped_data), headers)

    # Add a second sheet for percentages
    ws2 = wb.create_sheet(title="ECL Report (Percentages)")
    percentage_headers = [group_by_field, "% of 12 Month Reporting ECL", "% of Lifetime Reporting ECL", "% of Number of Accounts"]
    ws2.append(percentage_headers)

    # Ensure percentages data is not empty before adding rows
    if not df_percentages.empty:
        for row in dataframe_to_rows(df_percentages, index=False, header=False):
            ws2.append(row)

    # Add styling to the second sheet
    style_excel_sheet(ws2, len(grouped_data_percentages), percentage_headers)

    # Create an HTTP response with an Excel attachment
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="ecl summary report.xlsx"'

    # Save the workbook to the response
    wb.save(response)
    return response

def style_excel_sheet(ws, data_row_count, headers):
    """Apply styling to an Excel sheet."""
    header_fill = PatternFill(start_color="2d5c8e", end_color="2d5c8e", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    alignment = Alignment(horizontal="center", vertical="center")

    # Style the header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = alignment

    # Apply zebra striping for data rows
    light_fill = PatternFill(start_color="d1e7dd", end_color="d1e7dd", fill_type="solid")
    for row in ws.iter_rows(min_row=2, max_row=data_row_count + 1, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.fill = light_fill

    # Apply bold font for the grand total row if it exists
    if data_row_count > 0:
        for cell in ws[data_row_count + 1]:
            cell.font = Font(bold=True)

    # Adjust column widths
    for column in ws.columns:
        max_length = max(len(str(cell.value)) for cell in column if cell.value is not None)
        ws.column_dimensions[column[0].column_letter].width = max_length + 2




CSV_DIR = os.path.join(os.getcwd(), 'csv_files')
@login_required
@require_http_methods(["GET", "POST"])
def ecl_reconciliation_main_filter_view(request):
    # Initialize an empty errors dictionary
    errors = {}

    # Handle POST request (applying the filter)
    if request.method == 'POST':
        # Retrieve selected Reporting Dates and Run Keys from the form
        fic_mis_date1 = request.POST.get('fic_mis_date1')
        run_key1 = request.POST.get('run_key1')
        fic_mis_date2 = request.POST.get('fic_mis_date2')
        run_key2 = request.POST.get('run_key2')

        # Validate that both dates and run keys are provided
        if not fic_mis_date1:
            errors['fic_mis_date1'] = 'Please select Reporting Date 1.'
        if not run_key1:
            errors['run_key1'] = 'Please select Run Key 1.'
        if not fic_mis_date2:
            errors['fic_mis_date2'] = 'Please select Reporting Date 2.'
        if not run_key2:
            errors['run_key2'] = 'Please select Run Key 2.'

        # If there are no errors, proceed to filtering
        if not errors:
            # Store the selected filter values in session for later use
            request.session['fic_mis_date1'] = fic_mis_date1
            request.session['run_key1'] = run_key1
            request.session['fic_mis_date2'] = fic_mis_date2
            request.session['run_key2'] = run_key2

            # Retrieve filtered data based on the selected main filter values
            ecl_data1 = FCT_Reporting_Lines.objects.filter(fic_mis_date=fic_mis_date1, n_run_key=run_key1)
            ecl_data2 = FCT_Reporting_Lines.objects.filter(fic_mis_date=fic_mis_date2, n_run_key=run_key2)

            # Convert the filtered data into a DataFrame
            ecl_data1_df = pd.DataFrame(list(ecl_data1.values()))
            ecl_data2_df = pd.DataFrame(list(ecl_data2.values()))

            # Convert date fields to strings for both DataFrames
            for df in [ecl_data1_df, ecl_data2_df]:
                if 'fic_mis_date' in df.columns:
                    df['fic_mis_date'] = df['fic_mis_date'].astype(str)
                if 'd_maturity_date' in df.columns:
                    df['d_maturity_date'] = df['d_maturity_date'].astype(str)

            # Create the directory if it doesn't exist
            if not os.path.exists(CSV_DIR):
                os.makedirs(CSV_DIR)

            # Save the data as CSV files in the session (store the filenames)
            csv_filename1 = os.path.join(CSV_DIR, f"ecl_data_{fic_mis_date1}_{run_key1}.csv")
            csv_filename2 = os.path.join(CSV_DIR, f"ecl_data_{fic_mis_date2}_{run_key2}.csv")
            
            ecl_data1_df.to_csv(csv_filename1, index=False)
            ecl_data2_df.to_csv(csv_filename2, index=False)
            # Store the filenames in the session and explicitly mark the session as modified
            request.session['csv_filename1'] = csv_filename1
            request.session['csv_filename2'] = csv_filename2
            request.session.modified = True
            # Redirect to the sub-filter view
            return redirect('ecl_reconciliation_sub_filter_view')

    # Handle GET request to load Reporting Dates
    fic_mis_dates = FCT_Reporting_Lines.objects.order_by('-fic_mis_date').values_list('fic_mis_date', flat=True).distinct()

    # Check for AJAX requests to dynamically update run key dropdown
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        field_name = request.GET.get('field')
        field_value = request.GET.get('value')

        if field_name == 'fic_mis_date':
            # Fetch run keys corresponding to the selected FIC MIS date
            n_run_keys = FCT_Reporting_Lines.objects.filter(fic_mis_date=field_value).order_by('-n_run_key').values_list('n_run_key', flat=True).distinct()
            return JsonResponse({'n_run_keys': list(n_run_keys)})

    # If there are errors or it's a GET request, render the form with any errors
    return render(request, 'reports/ecl_reconciliation_main_filter.html', {
        'fic_mis_dates': fic_mis_dates,
        'errors': errors  # Pass the errors to the template
    })



@login_required
@require_http_methods(["GET", "POST"])
def ecl_reconciliation_sub_filter_view(request):
    # Retrieve the CSV filenames from the session
    csv_filename_1 = request.session.get('csv_filename1')
    csv_filename_2 = request.session.get('csv_filename2')

    # Check if the CSV files exist and display error if missing
    errors = []
    if not csv_filename_1:
        errors.append('The first dataset is missing. Please select a valid Reporting Date 1 and Run Key 1.')
    elif not os.path.exists(csv_filename_1):
        errors.append(f"The file for Reporting Date 1 does not exist: {csv_filename_1}")
    
    if not csv_filename_2:
        errors.append('The second dataset is missing. Please select a valid Reporting Date 2 and Run Key 2.')
    elif not os.path.exists(csv_filename_2):
        errors.append(f"The file for Reporting Date 2 does not exist: {csv_filename_2}")

    # If there are any errors, display them and do not proceed further
    if errors:
        for error in errors:
            messages.error(request, error)
        return render(request, 'reports/ecl_reconciliation_report_sub.html', {'errors': errors})

    # Load the data from the CSV files
    ecl_data_df_1 = pd.read_csv(csv_filename_1)
    ecl_data_df_2 = pd.read_csv(csv_filename_2)

    # Determine which dataframe has the higher and lower date based on fic_mis_date
    if ecl_data_df_1['fic_mis_date'].max() > ecl_data_df_2['fic_mis_date'].max():
        df_higher_date = ecl_data_df_1
        df_lower_date = ecl_data_df_2
    else:
        df_higher_date = ecl_data_df_2
        df_lower_date = ecl_data_df_1

    # Merge based on fic_mis_date and other keys
    merged_data = pd.merge(
        df_higher_date, 
        df_lower_date, 
        on=['fic_mis_date', 'n_run_key', 'n_stage_descr','v_ccy_code','n_prod_segment','n_prod_type','n_loan_type'],  
        how='outer', 
        suffixes=('_higher', '_lower')
    )

    # Fill NaN values with zeros for numeric fields
    merged_data.fillna({
        'n_exposure_at_default_ncy_higher': 0, 'n_exposure_at_default_ncy_lower': 0,
        'n_exposure_at_default_rcy_higher': 0, 'n_exposure_at_default_rcy_lower': 0,
        'n_12m_ecl_rcy_higher': 0, 'n_12m_ecl_rcy_lower': 0,
        'n_lifetime_ecl_rcy_higher': 0, 'n_lifetime_ecl_rcy_lower': 0
    }, inplace=True)

    # Apply sub-filters to the merged data
    v_ccy_code = request.GET.get('v_ccy_code')
    n_prod_segment = request.GET.get('n_prod_segment')
    n_prod_type = request.GET.get('n_prod_type')
    n_stage_descr = request.GET.get('n_stage_descr')
    n_loan_type = request.GET.get('n_loan_type')

    if v_ccy_code:
        merged_data = merged_data[merged_data['v_ccy_code'] == v_ccy_code]
    if n_prod_segment:
        merged_data = merged_data[merged_data['n_prod_segment'] == n_prod_segment]
    if n_prod_type:
        merged_data = merged_data[merged_data['n_prod_type'] == n_prod_type]
    if n_stage_descr:
        merged_data = merged_data[merged_data['n_stage_descr'] == n_stage_descr]
    if n_loan_type:
        merged_data = merged_data[merged_data['n_loan_type'] == n_loan_type]

    # Group by functionality after merging and applying filters
    group_by_field = request.GET.get('group_by_field', 'n_stage_descr') 

    # Group data and calculate the sum, while counting unique account numbers for higher and lower datasets
    grouped_data = merged_data.groupby(group_by_field).agg({
        'n_exposure_at_default_ncy_higher': 'sum',
        'n_exposure_at_default_ncy_lower': 'sum',
        'n_exposure_at_default_rcy_higher': 'sum',
        'n_exposure_at_default_rcy_lower': 'sum',
        'n_12m_ecl_rcy_higher': 'sum',
        'n_12m_ecl_rcy_lower': 'sum',
        'n_lifetime_ecl_rcy_higher': 'sum',
        'n_lifetime_ecl_rcy_lower': 'sum',
        # Counting distinct account numbers for higher and lower
        'n_account_number_higher': pd.Series.nunique,  
        'n_account_number_lower': pd.Series.nunique    
    }).reset_index()

    # Create new columns for account counts based on the 'n_account_number' column
    grouped_data['n_accounts_in_higher'] = grouped_data['n_account_number_higher']
    grouped_data['n_accounts_in_lower'] = grouped_data['n_account_number_lower']

    # Calculate differences (higher date minus lower date)
    grouped_data['difference_ead_ncy'] = grouped_data['n_exposure_at_default_ncy_higher'] - grouped_data['n_exposure_at_default_ncy_lower']
    grouped_data['difference_ead_rcy'] = grouped_data['n_exposure_at_default_rcy_higher'] - grouped_data['n_exposure_at_default_rcy_lower']
    grouped_data['difference_12m_ecl'] = grouped_data['n_12m_ecl_rcy_higher'] - grouped_data['n_12m_ecl_rcy_lower']
    grouped_data['difference_lifetime_ecl'] = grouped_data['n_lifetime_ecl_rcy_higher'] - grouped_data['n_lifetime_ecl_rcy_lower']

    # Status calculation based on 12-month ECL differences
    grouped_data['status_ead_ncy'] = grouped_data['difference_12m_ecl'].apply(
        lambda x: 'Increased' if x > 0 else ('Decreased' if x < 0 else 'No Change')
    )

    # Convert numeric columns to Python native types for JSON serialization
    grouped_data = grouped_data.applymap(lambda x: int(x) if isinstance(x, np.int64) else x)
    grouped_data = grouped_data.applymap(lambda x: float(x) if isinstance(x, np.float64) else x)

    # Convert to JSON-serializable format (a list of dictionaries) to store in session
    grouped_data_json = grouped_data.to_dict(orient='records')

    # Grand totals for columns and differences, also converted to Python native types
    grand_totals = {
        'n_exposure_at_default_ncy_higher': int(grouped_data['n_exposure_at_default_ncy_higher'].sum()),
        'n_exposure_at_default_rcy_higher': int(grouped_data['n_exposure_at_default_rcy_higher'].sum()),
        'n_12m_ecl_rcy_higher': int(grouped_data['n_12m_ecl_rcy_higher'].sum()),
        'n_lifetime_ecl_rcy_higher': int(grouped_data['n_lifetime_ecl_rcy_higher'].sum()),
        'n_exposure_at_default_ncy_lower': int(grouped_data['n_exposure_at_default_ncy_lower'].sum()),
        'n_exposure_at_default_rcy_lower': int(grouped_data['n_exposure_at_default_rcy_lower'].sum()),
        'n_12m_ecl_rcy_lower': int(grouped_data['n_12m_ecl_rcy_lower'].sum()),
        'n_lifetime_ecl_rcy_lower': int(grouped_data['n_lifetime_ecl_rcy_lower'].sum()),
        'difference_ead_rcy': int(grouped_data['difference_ead_rcy'].sum()),
        'difference_12m_ecl': int(grouped_data['difference_12m_ecl'].sum()),
        'difference_lifetime_ecl': int(grouped_data['difference_lifetime_ecl'].sum()),
        'n_accounts_in_higher': int(grouped_data['n_accounts_in_higher'].sum()),
        'n_accounts_in_lower': int(grouped_data['n_accounts_in_lower'].sum()),
    }
    # Calculate percentages for 12 Month ECL, Lifetime ECL, differences, and total accounts
    grouped_data['percent_12m_ecl_higher'] = (grouped_data['n_12m_ecl_rcy_higher'] / grand_totals['n_12m_ecl_rcy_higher'] * 100) if grand_totals['n_12m_ecl_rcy_higher'] else 0
    grouped_data['percent_12m_ecl_lower'] = (grouped_data['n_12m_ecl_rcy_lower'] / grand_totals['n_12m_ecl_rcy_lower'] * 100) if grand_totals['n_12m_ecl_rcy_lower'] else 0
    grouped_data['percent_lifetime_ecl_higher'] = (grouped_data['n_lifetime_ecl_rcy_higher'] / grand_totals['n_lifetime_ecl_rcy_higher'] * 100) if grand_totals['n_lifetime_ecl_rcy_higher'] else 0
    grouped_data['percent_lifetime_ecl_lower'] = (grouped_data['n_lifetime_ecl_rcy_lower'] / grand_totals['n_lifetime_ecl_rcy_lower'] * 100) if grand_totals['n_lifetime_ecl_rcy_lower'] else 0
    grouped_data['percent_total_accounts_higher'] = (grouped_data['n_accounts_in_higher'] / grand_totals['n_accounts_in_higher'] * 100) if grand_totals['n_accounts_in_higher'] else 0
    grouped_data['percent_total_accounts_lower'] = (grouped_data['n_accounts_in_lower'] / grand_totals['n_accounts_in_lower'] * 100) if grand_totals['n_accounts_in_lower'] else 0
    # Calculate percentage differences based on higher values
    grouped_data['percent_difference_12m_ecl'] = (grouped_data['difference_12m_ecl'] / grouped_data['n_12m_ecl_rcy_higher'] * 100).replace([np.inf, -np.inf], 0).fillna(0)
    grouped_data['percent_difference_lifetime_ecl'] = (grouped_data['difference_lifetime_ecl'] / grouped_data['n_lifetime_ecl_rcy_higher'] * 100).replace([np.inf, -np.inf], 0).fillna(0)
    grouped_data['percent_difference_ead_rcy'] = (grouped_data['difference_ead_rcy'] / grouped_data['n_exposure_at_default_rcy_higher'] * 100).replace([np.inf, -np.inf], 0).fillna(0)


    # Convert grouped_data to JSON-serializable format for percentages
    grouped_data_percentages = grouped_data.to_dict(orient='records')

    # Store the grouped data and grand totals in the session for Excel export (JSON serializable format)
    request.session['grouped_data'] = grouped_data_json
    request.session['group_by_field'] = group_by_field
    request.session['grand_totals'] = grand_totals
    request.session['grouped_data_percentages'] = grouped_data_percentages

    # Distinct values for sub-filters
    distinct_currency_codes = list(merged_data['v_ccy_code'].unique())
    distinct_prod_segments = list(merged_data['n_prod_segment'].unique())
    distinct_prod_types = list(merged_data['n_prod_type'].unique())
    distinct_stage_descrs = list(merged_data['n_stage_descr'].unique())
    distinct_loan_types = list(merged_data['n_loan_type'].unique())

    # Render the sub-filter view with the reconciliation data
    return render(request, 'reports/ecl_reconciliation_report_sub.html', {
        'grouped_data': grouped_data_json,
        'grouped_data_percentages': grouped_data_percentages,
        'group_by_field': group_by_field,
        'distinct_currency_codes': distinct_currency_codes,
        'distinct_prod_segments': distinct_prod_segments,
        'distinct_prod_types': distinct_prod_types,
        'distinct_stage_descrs': distinct_stage_descrs,
        'distinct_loan_types': distinct_loan_types,
        'grand_totals': grand_totals,
    })


@login_required
@require_http_methods(["POST"])
def export_ecl_reconciliation_to_excel(request):
    # Retrieve the filtered data, grouped percentages, and grand totals from session
    grouped_data = request.session.get('grouped_data', [])
    grouped_data_percentages = request.session.get('grouped_data_percentages', [])
    group_by_field = request.session.get('group_by_field', 'n_stage_descr')
    grand_totals = request.session.get('grand_totals', {})
    fic_mis_date1 = request.session.get('fic_mis_date1', '')
    run_key1 = request.session.get('run_key1', '')
    fic_mis_date2 = request.session.get('fic_mis_date2', '')
    run_key2 = request.session.get('run_key2', '')

    # Create a new Excel workbook
    wb = Workbook()

    # First Sheet: ECL Reconciliation Report (Absolute Values)
    ws1 = wb.active
    ws1.title = "ECL Reconciliation Report"

    # Merge cells for the headers
    ws1.merge_cells('B1:E1')  # Merge columns B to E in the first row for Date 1
    ws1['B1'] = f"Reporting Date 1 ({fic_mis_date1}) - Run Key 1 ({run_key1})"
    ws1['B1'].alignment = Alignment(horizontal='center', vertical='center')

    ws1.merge_cells('F1:I1')  # Merge columns F to I in the first row for Date 2
    ws1['F1'] = f"Reporting Date 2 ({fic_mis_date2}) - Run Key 2 ({run_key2})"
    ws1['F1'].alignment = Alignment(horizontal='center', vertical='center')

    ws1.merge_cells('J1:L1')  # Merge columns J to L in the first row for Differences
    ws1['J1'] = "Differences"
    ws1['J1'].alignment = Alignment(horizontal='center', vertical='center')

    ws1.merge_cells('M1:N1')  # Merge columns M to N in the first row for Total Accounts
    ws1['M1'] = "Total Accounts"
    ws1['M1'].alignment = Alignment(horizontal='center', vertical='center')


    # Add the main headers
    headers = [
        group_by_field,
        f"Reporting Date 1 ({fic_mis_date1}) - Run Key 1 ({run_key1})",
        "",
        "",
        "",
        f"Reporting Date 2 ({fic_mis_date2}) - Run Key 2 ({run_key2})",
        "",
        "",
        "",
        "Differences",
        "",
        "",
        "Total Accounts",
        ""
    ]

    # Add sub-headers for the columns under the merged headers
    sub_headers = [
        group_by_field,
        "EAD Orig Currency",
        "EAD Reporting Currency",
        "12 Month ECL",
        "Lifetime ECL",
        "EAD Orig Currency",
        "EAD Reporting Currency",
        "12 Month ECL",
        "Lifetime ECL",
        "EAD Reporting Currency Difference",
        "12 Month ECL Difference",
        "Lifetime ECL Difference",
        f"Total Accounts ({fic_mis_date1})",
        f"Total Accounts ({fic_mis_date2})"
    ]
    ws1.append(sub_headers)

    # Add the grouped data to the first sheet
    for row in grouped_data:
        ws1.append([
            row.get(group_by_field, ''),
            row.get('n_exposure_at_default_ncy_higher', 0),
            row.get('n_exposure_at_default_rcy_higher', 0),
            row.get('n_12m_ecl_rcy_higher', 0),
            row.get('n_lifetime_ecl_rcy_higher', 0),
            row.get('n_exposure_at_default_ncy_lower', 0),
            row.get('n_exposure_at_default_rcy_lower', 0),
            row.get('n_12m_ecl_rcy_lower', 0),
            row.get('n_lifetime_ecl_rcy_lower', 0),
            row.get('difference_ead_rcy', 0),
            row.get('difference_12m_ecl', 0),
            row.get('difference_lifetime_ecl', 0),
            row.get('n_accounts_in_higher', 0),
            row.get('n_accounts_in_lower', 0),
        ])

    # Add the grand totals to the first sheet
    ws1.append([
        'Grand Total',
        grand_totals.get('n_exposure_at_default_ncy_higher', 0),
        grand_totals.get('n_exposure_at_default_rcy_higher', 0),
        grand_totals.get('n_12m_ecl_rcy_higher', 0),
        grand_totals.get('n_lifetime_ecl_rcy_higher', 0),
        grand_totals.get('n_exposure_at_default_ncy_lower', 0),
        grand_totals.get('n_exposure_at_default_rcy_lower', 0),
        grand_totals.get('n_12m_ecl_rcy_lower', 0),
        grand_totals.get('n_lifetime_ecl_rcy_lower', 0),
        grand_totals.get('difference_ead_rcy', 0),
        grand_totals.get('difference_12m_ecl', 0),
        grand_totals.get('difference_lifetime_ecl', 0),
        grand_totals.get('n_accounts_in_higher', 0),
        grand_totals.get('n_accounts_in_lower', 0),
    ])

    # Second Sheet: Percentage Table
    ws2 = wb.create_sheet(title="ECL Percentages")

     # Merge cells for the headers
    ws2.merge_cells('B1:C1')  # Merge columns B to E in the first row for Date 1
    ws2['B1'] = f"Reporting Date 1 ({fic_mis_date1}) - Run Key 1 ({run_key1})"
    ws2['B1'].alignment = Alignment(horizontal='center', vertical='center')

    ws2.merge_cells('D1:E1')  # Merge columns F to I in the first row for Date 2
    ws2['D1'] = f"Reporting Date 2 ({fic_mis_date2}) - Run Key 2 ({run_key2})"
    ws2['D1'].alignment = Alignment(horizontal='center', vertical='center')

    ws2.merge_cells('F1:G1')  # Merge columns J to L in the first row for Differences
    ws2['F1'] = "Total Accounts"
    ws2['F1'].alignment = Alignment(horizontal='center', vertical='center')

    ws2.merge_cells('H1:J1')  # Merge columns M to N in the first row for Total Accounts
    ws2['H1'] = "Differences"
    ws2['H1'].alignment = Alignment(horizontal='center', vertical='center')

    # Add headers for percentages
    percentage_headers = [
        group_by_field,
        "% of 12 Month ECL (Date 1)",
        "% of Lifetime ECL (Date 1)",
        "% of 12 Month ECL (Date 2)",
        "% of Lifetime ECL (Date 2)",
        "% of Total Accounts (Date 1)",
        "% of Total Accounts (Date 2)",
        "% of EAD Reporting Currency Difference",
        "% of 12 Month ECL Difference",
        "% of Lifetime ECL Difference",
    ]
    ws2.append(percentage_headers)

    # Add grouped percentage data to the second sheet
    for row in grouped_data_percentages:
        ws2.append([
            row.get(group_by_field, ''),
            row.get('percent_12m_ecl_higher', 0),
            row.get('percent_lifetime_ecl_higher', 0),
            row.get('percent_12m_ecl_lower', 0),
            row.get('percent_lifetime_ecl_lower', 0),
            row.get('percent_total_accounts_higher', 0),
            row.get('percent_total_accounts_lower', 0),
            row.get('percent_difference_ead_rcy', 0),
            row.get('percent_difference_12m_ecl', 0),
            row.get('percent_difference_lifetime_ecl', 0),
        ])

    # Apply styling to both sheets (optional: customize as needed)
    style_excel_sheet(ws1, len(grouped_data), sub_headers)
    style_excel_sheet(ws2, len(grouped_data_percentages), percentage_headers)

    # Create an HTTP response with an Excel attachment
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="ecl_reconciliation_{fic_mis_date1}_{fic_mis_date2}.xlsx"'

    # Save the workbook to the response
    wb.save(response)
    return response


def style_excel_sheet(ws, data_row_count, headers):
    """Applies consistent styling to the Excel sheet."""
    header_fill = PatternFill(start_color="2d5c8e", end_color="2d5c8e", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    alignment = Alignment(horizontal="center", vertical="center")

    # Style the header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = alignment

    # Style the sub-headers
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = alignment

    # Zebra striping for data rows
    light_fill = PatternFill(start_color="d1e7dd", end_color="d1e7dd", fill_type="solid")
    for row in ws.iter_rows(min_row=2, max_row=data_row_count + 3, min_col=1, max_col=len(headers)):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left")
            cell.fill = light_fill

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter if not isinstance(col[0], openpyxl.cell.cell.MergedCell) else None
        if column_letter:  # Only adjust if column is valid (not a merged cell)
            for cell in col:
                if cell.value and not isinstance(cell, openpyxl.cell.cell.MergedCell):  # Skip merged cells
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

CSV_DIR = os.path.join(os.getcwd(), 'csv_files')

@login_required
@require_http_methods(["GET", "POST"])
def ecl_account_reconciliation_main_filter_view(request):
    errors = {}

    # Handle POST request (applying the filter)
    if request.method == 'POST':
        # Retrieve selected Reporting Dates and Run Keys from the form
        fic_mis_date1 = request.POST.get('fic_mis_date1')
        run_key1 = request.POST.get('run_key1')
        fic_mis_date2 = request.POST.get('fic_mis_date2')
        run_key2 = request.POST.get('run_key2')

        # Validate that both dates and run keys are provided
        if not fic_mis_date1:
            errors['fic_mis_date1'] = 'Please select Reporting Date 1.'
        if not run_key1:
            errors['run_key1'] = 'Please select Run Key 1.'
        if not fic_mis_date2:
            errors['fic_mis_date2'] = 'Please select Reporting Date 2.'
        if not run_key2:
            errors['run_key2'] = 'Please select Run Key 2.'

        # If no errors, proceed with filter
        if not errors:
            request.session['fic_mis_date1'] = fic_mis_date1
            request.session['run_key1'] = run_key1
            request.session['fic_mis_date2'] = fic_mis_date2
            request.session['run_key2'] = run_key2

            # Retrieve filtered data based on the selected main filter values
            ecl_data1 = FCT_Reporting_Lines.objects.filter(fic_mis_date=fic_mis_date1, n_run_key=run_key1)
            ecl_data2 = FCT_Reporting_Lines.objects.filter(fic_mis_date=fic_mis_date2, n_run_key=run_key2)

            # Convert the filtered data into a DataFrame
            ecl_data1_df = pd.DataFrame(list(ecl_data1.values()))
            ecl_data2_df = pd.DataFrame(list(ecl_data2.values()))

            # Ensure both DataFrames use the same set of account numbers
            common_account_numbers = set(ecl_data1_df['n_account_number']).intersection(ecl_data2_df['n_account_number'])
            ecl_data1_df = ecl_data1_df[ecl_data1_df['n_account_number'].isin(common_account_numbers)]
            ecl_data2_df = ecl_data2_df[ecl_data2_df['n_account_number'].isin(common_account_numbers)]

            # Convert date fields to strings for both DataFrames
            for df in [ecl_data1_df, ecl_data2_df]:
                if 'fic_mis_date' in df.columns:
                    df['fic_mis_date'] = df['fic_mis_date'].astype(str)
                if 'd_maturity_date' in df.columns:
                    df['d_maturity_date'] = df['d_maturity_date'].astype(str)

            # Create the directory if it doesn't exist
            if not os.path.exists(CSV_DIR):
                os.makedirs(CSV_DIR)

            # Save the data as CSV files in the session
            csv_filename1 = os.path.join(CSV_DIR, f"ecl_data_{fic_mis_date1}_{run_key1}.csv")
            csv_filename2 = os.path.join(CSV_DIR, f"ecl_data_{fic_mis_date2}_{run_key2}.csv")
            ecl_data1_df.to_csv(csv_filename1, index=False)
            ecl_data2_df.to_csv(csv_filename2, index=False)
            request.session['csv_filename1'] = csv_filename1
            request.session['csv_filename2'] = csv_filename2
            request.session.modified = True

            # Redirect to the sub-filter view
            return redirect('ecl_account_reconciliation_sub_filter')

    # Handle GET request to load Reporting Dates
    fic_mis_dates = FCT_Reporting_Lines.objects.order_by('-fic_mis_date').values_list('fic_mis_date', flat=True).distinct()

    # AJAX request for dynamically updating the Run Key dropdowns
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        field_name = request.GET.get('field')
        field_value = request.GET.get('value')

        if field_name == 'fic_mis_date':
            # Fetch run keys corresponding to the selected FIC MIS date
            n_run_keys = FCT_Reporting_Lines.objects.filter(fic_mis_date=field_value).order_by('-n_run_key').values_list('n_run_key', flat=True).distinct()
            return JsonResponse({'n_run_keys': list(n_run_keys)})

    # If there are errors or it's a GET request, render the form with any errors
    return render(request, 'reports/ecl_account_reconciliation_main_filter.html', {
        'fic_mis_dates': fic_mis_dates,
        'errors': errors  # Pass the errors to the template
    })


@login_required
@require_http_methods(["GET", "POST"])
def ecl_account_reconciliation_sub_filter_view(request):
    # Retrieve CSV filenames from the session
    csv_filename_1 = request.session.get('csv_filename1')
    csv_filename_2 = request.session.get('csv_filename2')

    # Load CSV files into DataFrames
    ecl_data_df_1 = pd.read_csv(csv_filename_1)
    ecl_data_df_2 = pd.read_csv(csv_filename_2)

    # Merge data based on account number using an inner join
    merged_data = pd.merge(
        ecl_data_df_1,
        ecl_data_df_2,
        on=['n_account_number'],
        how='inner',
        suffixes=('_prev', '_curr')
    )

    # Filter based on account number from the request
    account_number = request.GET.get('n_account_number')
    if account_number:
        merged_data = merged_data[merged_data['n_account_number'] == account_number]

    # Pass the merged data to the template (handle single account selection)
    account_details = {}
    if not merged_data.empty:
        account_details = {
            'date_prev': merged_data['fic_mis_date_prev'].iloc[0],
            'date_curr': merged_data['fic_mis_date_curr'].iloc[0],
            'n_account_number': merged_data['n_account_number'].iloc[0],
            'n_run_key_prev': merged_data['n_run_key_prev'].iloc[0],
            'n_run_key_curr': merged_data['n_run_key_curr'].iloc[0],
            'v_ccy_code_prev': merged_data['v_ccy_code_prev'].iloc[0],
            'v_ccy_code_curr': merged_data['v_ccy_code_curr'].iloc[0],
            'balance_outstanding_prev': merged_data['n_carrying_amount_rcy_prev'].iloc[0],
            'balance_outstanding_curr': merged_data['n_carrying_amount_rcy_curr'].iloc[0],
            'exposure_at_default_prev': merged_data['n_exposure_at_default_rcy_prev'].iloc[0],
            'exposure_at_default_curr': merged_data['n_exposure_at_default_rcy_curr'].iloc[0],
            'ifrs_stage_prev': merged_data['n_stage_descr_prev'].iloc[0],
            'ifrs_stage_curr': merged_data['n_stage_descr_curr'].iloc[0],
            'delq_band_code_prev': merged_data['n_delq_band_code_prev'].iloc[0],
            'delq_band_code_curr': merged_data['n_delq_band_code_curr'].iloc[0],
            'maturity_date_prev': merged_data['d_maturity_date_prev'].iloc[0],
            'maturity_date_curr': merged_data['d_maturity_date_curr'].iloc[0],
            'twelve_month_pd_prev': merged_data['n_twelve_months_pd_prev'].iloc[0],
            'twelve_month_pd_curr': merged_data['n_twelve_months_pd_curr'].iloc[0],
            'lifetime_pd_prev': merged_data['n_lifetime_pd_prev'].iloc[0],
            'lifetime_pd_curr': merged_data['n_lifetime_pd_curr'].iloc[0],
            'lgd_prev': merged_data['n_lgd_percent_prev'].iloc[0],
            'lgd_curr': merged_data['n_lgd_percent_curr'].iloc[0],
            'twelve_month_ecl_prev': merged_data['n_12m_ecl_rcy_prev'].iloc[0],  # Correct column
            'twelve_month_ecl_curr': merged_data['n_12m_ecl_rcy_curr'].iloc[0],  # Correct column
            'lifetime_ecl_prev': merged_data['n_lifetime_ecl_rcy_prev'].iloc[0],  # Correct column
            'lifetime_ecl_curr': merged_data['n_lifetime_ecl_rcy_curr'].iloc[0],  # Correct column
            'prod_segment_prev': merged_data['n_prod_segment_prev'].iloc[0],
            'prod_segment_curr': merged_data['n_prod_segment_curr'].iloc[0],
        }

    # Get account options for the dropdown
    account_options = merged_data['n_account_number'].unique().tolist()

    return render(request, 'reports/ecl_account_reconciliation_report.html', {
        'account_details': account_details,
        'account_options': account_options,
        'selected_account': account_number
    })



@login_required
@require_http_methods(["GET"])
def export_full_ecl_report_to_excel(request):
    # Retrieve CSV filenames from the session
    csv_filename_1 = request.session.get('csv_filename1')
    csv_filename_2 = request.session.get('csv_filename2')

    # Load CSV files into DataFrames
    ecl_data_df_1 = pd.read_csv(csv_filename_1)
    ecl_data_df_2 = pd.read_csv(csv_filename_2)

    # Merge data based on account number using an inner join
    merged_data = pd.merge(
        ecl_data_df_1,
        ecl_data_df_2,
        on=['n_account_number'],
        how='inner',
        suffixes=('_prev', '_curr')
    )

    # Get the `fic_mis_date` and `n_run_key` for the two periods
    fic_mis_date_1 = merged_data['fic_mis_date_prev'].iloc[0]
    fic_mis_date_2 = merged_data['fic_mis_date_curr'].iloc[0]
    run_key_1 = merged_data['n_run_key_prev'].iloc[0]
    run_key_2 = merged_data['n_run_key_curr'].iloc[0]

    # Create a new Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Full ECL Reconciliation Report"

    # Colors for different headers
    header_color_1 = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")  # Yellow
    header_color_2 = PatternFill(start_color="A9D08E", end_color="A9D08E", fill_type="solid")  # Green

    # Add column headers for the Excel sheet using fic_mis_date and n_run_key
    headers = [
        'Account ID',
        'Product Segment',
        'Currency Code',
        f'Balance Outstanding ({fic_mis_date_1} - Run Key {run_key_1})',
        f'Balance Outstanding ({fic_mis_date_2} - Run Key {run_key_2})',
        f'Exposure at Default ({fic_mis_date_1} - Run Key {run_key_1})',
        f'Exposure at Default ({fic_mis_date_2} - Run Key {run_key_2})',
        f'IFRS Stage ({fic_mis_date_1} - Run Key {run_key_1})',
        f'IFRS Stage ({fic_mis_date_2} - Run Key {run_key_2})',
        f'Delinquency Band Code ({fic_mis_date_1} - Run Key {run_key_1})',
        f'Delinquency Band Code ({fic_mis_date_2} - Run Key {run_key_2})',
        f'Maturity Date ({fic_mis_date_1} - Run Key {run_key_1})',
        f'Maturity Date ({fic_mis_date_2} - Run Key {run_key_2})',
        f'12 Month PD ({fic_mis_date_1} - Run Key {run_key_1})',
        f'12 Month PD ({fic_mis_date_2} - Run Key {run_key_2})',
        f'Lifetime PD ({fic_mis_date_1} - Run Key {run_key_1})',
        f'Lifetime PD ({fic_mis_date_2} - Run Key {run_key_2})',
        f'LGD ({fic_mis_date_1} - Run Key {run_key_1})',
        f'LGD ({fic_mis_date_2} - Run Key {run_key_2})',
        f'12 Month Reporting ECL ({fic_mis_date_1} - Run Key {run_key_1})',
        f'12 Month Reporting ECL ({fic_mis_date_2} - Run Key {run_key_2})',
        f'Lifetime Reporting ECL ({fic_mis_date_1} - Run Key {run_key_1})',
        f'Lifetime Reporting ECL ({fic_mis_date_2} - Run Key {run_key_2})'
    ]
    
    ws.append(headers)

    # Apply color formatting for the headers
    for col_num, cell in enumerate(ws[1], start=1):
        if col_num >= 4 and col_num % 2 == 0:  # Columns related to fic_mis_date_2
            cell.fill = header_color_2
        elif col_num >= 4:  # Columns related to fic_mis_date_1
            cell.fill = header_color_1

    # Add the rows of data to the Excel sheet
    for index, row in merged_data.iterrows():
        ws.append([
            row['n_account_number'],
            row.get('n_prod_segment_prev', ''),
            row.get('v_ccy_code_prev', ''),
            row.get('n_carrying_amount_rcy_prev', ''), row.get('n_carrying_amount_rcy_curr', ''),
            row.get('n_exposure_at_default_rcy_prev', ''), row.get('n_exposure_at_default_rcy_curr', ''),
            row.get('n_stage_descr_prev', ''), row.get('n_stage_descr_curr', ''),
            row.get('n_delq_band_code_prev', ''), row.get('n_delq_band_code_curr', ''),
            row.get('d_maturity_date_prev', ''), row.get('d_maturity_date_curr', ''),
            row.get('n_twelve_months_pd_prev', ''), row.get('n_twelve_months_pd_curr', ''),
            row.get('n_lifetime_pd_prev', ''), row.get('n_lifetime_pd_curr', ''),
            row.get('n_lgd_percent_prev', ''), row.get('n_lgd_percent_curr', ''),
            row.get('n_12m_ecl_rcy_prev', ''), row.get('n_12m_ecl_rcy_curr', ''),
            row.get('n_lifetime_ecl_rcy_prev', ''), row.get('n_lifetime_ecl_rcy_curr', ''),
        ])

    # Set up the response as an Excel file download
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="ecl_full_report_account_reconcilation.xlsx"'

    # Save the workbook to the response
    wb.save(response)

    return response


@login_required
@require_http_methods(["GET", "POST"])
def pd_analysis_main_filter_view(request):
    # Handle POST request (applying the filter)
    if request.method == 'POST':
        # Retrieve selected Reporting Date and Run Key from the form
        fic_mis_date = request.POST.get('fic_mis_date')
        n_run_key = request.POST.get('n_run_key')
        
        if fic_mis_date and n_run_key:
            # Store the selected filter values in session for later use
            request.session['fic_mis_date'] = fic_mis_date
            request.session['n_run_key'] = n_run_key

            # Retrieve filtered data based on the selected main filter values
            pd_data = FCT_Reporting_Lines.objects.filter(fic_mis_date=fic_mis_date, n_run_key=n_run_key)

            # Convert the filtered data into a DataFrame
            pd_data_df = pd.DataFrame(list(pd_data.values()))

            # Convert date fields to strings
            if 'fic_mis_date' in pd_data_df.columns:
                pd_data_df['fic_mis_date'] = pd_data_df['fic_mis_date'].astype(str)

            if 'd_maturity_date' in pd_data_df.columns:
                pd_data_df['d_maturity_date'] = pd_data_df['d_maturity_date'].astype(str)

            # Save the data as a CSV file in the session (store the filename)
            csv_filename = os.path.join(CSV_DIR, f"pd_data_{fic_mis_date}_{n_run_key}.csv")
            pd_data_df.to_csv(csv_filename, index=False)
            request.session['csv_filename'] = csv_filename

            # Redirect to the sub-filter view
            return redirect('pd_analysis_sub_filter_view')

    # Handle GET request for Reporting Dates
    fic_mis_dates = FCT_Reporting_Lines.objects.order_by('-fic_mis_date').values_list('fic_mis_date', flat=True).distinct()

    # AJAX request handling for dynamic run key dropdown
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        field_name = request.GET.get('field')
        field_value = request.GET.get('value')

        if field_name == 'fic_mis_date':
            # Fetch run keys corresponding to the selected FIC MIS date
            n_run_keys = FCT_Reporting_Lines.objects.filter(fic_mis_date=field_value).order_by('-n_run_key').values_list('n_run_key', flat=True).distinct()
            return JsonResponse({'n_run_keys': list(n_run_keys)})

    # Render the main filter template with Reporting Dates
    return render(request, 'reports/pd_analysis_main_filter.html', {'fic_mis_dates': fic_mis_dates})


@login_required
@require_http_methods(["GET", "POST"])
def pd_analysis_sub_filter_view(request):
    # Retrieve the CSV file from the session
    csv_filename = request.session.get('csv_filename')

    # Handle no data scenario
    if not csv_filename or not os.path.exists(csv_filename):
        error_message = 'No data available. Please apply the main filter.'
        return render(request, 'reports/pd_analysis_report.html', {'error_message': error_message})

    # Load the CSV data into a DataFrame
    pd_data_df = pd.read_csv(csv_filename)

    # Apply filters
    n_prod_segment = request.GET.get('n_prod_segment')
    n_prod_type = request.GET.get('n_prod_type')
    n_stage_descr = request.GET.get('n_stage_descr')
    n_loan_type = request.GET.get('n_loan_type')

    # Filter the data dynamically
    if n_prod_segment:
        pd_data_df = pd_data_df[pd_data_df['n_prod_segment'] == n_prod_segment]
    if n_prod_type:
        pd_data_df = pd_data_df[pd_data_df['n_prod_type'] == n_prod_type]
    if n_stage_descr:
        pd_data_df = pd_data_df[pd_data_df['n_stage_descr'] == n_stage_descr]
    if n_loan_type:
        pd_data_df = pd_data_df[pd_data_df['n_loan_type'] == n_loan_type]

    # Retrieve the selected group by field
    group_by_field = request.GET.get('group_by_field', 'n_stage_descr')

    # Group the data by selected field and calculate the PD ranges
    grouped_data = pd_data_df.groupby(group_by_field).agg({
        'n_twelve_months_pd': ['min', 'max'],
        'n_lifetime_pd': ['min', 'max'],
        'n_lgd_percent': ['min', 'max'],
    }).reset_index()

    # Prepare columns for display
    grouped_data.columns = [
        group_by_field, '12 Month PD Min', '12 Month PD Max', 'Lifetime PD Min', 'Lifetime PD Max', 'LGD Min', 'LGD Max'
    ]
    
    grouped_data_list = grouped_data.to_dict(orient='records')

    # Store the grouped data and group_by_field in the session for download
    request.session['grouped_data'] = grouped_data_list
    request.session['group_by_field'] = group_by_field


    # Get distinct values for filters (for dropdowns)
    distinct_prod_segments = pd_data_df['n_prod_segment'].unique()
    distinct_prod_types = pd_data_df['n_prod_type'].unique()
    distinct_stage_descrs = pd_data_df['n_stage_descr'].unique()
    distinct_loan_types = pd_data_df['n_loan_type'].unique()

    # Render the template
    return render(request, 'reports/pd_analysis_report.html', {
        'grouped_data': grouped_data_list,
        'group_by_field': group_by_field,
        'distinct_prod_segments': distinct_prod_segments,
        'distinct_prod_types': distinct_prod_types,
        'distinct_stage_descrs': distinct_stage_descrs,
        'distinct_loan_types': distinct_loan_types,
        'selected_prod_segment': n_prod_segment,
        'selected_prod_type': n_prod_type,
        'selected_stage_descr': n_stage_descr,
        'selected_loan_type': n_loan_type,
        'error_message': None
    })


@login_required
@require_http_methods(["POST"])
def export_pd_report_to_excel(request):
    # Retrieve the grouped data and group by field from the session
    grouped_data = request.session.get('grouped_data', [])
    group_by_field = request.session.get('group_by_field', 'n_pd_term_structure_name')

    # Create a new Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "PD Analysis Report"

    # Define the headers for the Excel sheet
    headers = [
        group_by_field,
        "12 Month PD Min",
        "12 Month PD Max",
        "Lifetime PD Min",
        "Lifetime PD Max",
        "LGD Min",
        "LGD Max"
    ]
    ws.append(headers)

    # Add the grouped data to the Excel sheet
    for row in grouped_data:
        ws.append([
            row.get(group_by_field, ''),
            row.get('12 Month PD Min', ''),
            row.get('12 Month PD Max', ''),
            row.get('Lifetime PD Min', ''),
            row.get('Lifetime PD Max', ''),
            row.get('LGD Min', ''),
            row.get('LGD Max', '')
        ])

    # Set up the HTTP response as an Excel file download
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="pd_analysis_report.xlsx"'

    # Save the workbook to the response
    wb.save(response)

    return response



@login_required
@require_http_methods(["GET", "POST"])
def water_fall_main_filter_view(request):
    errors = {}

    # Handle POST request (applying the filter)
    if request.method == 'POST':
        # Retrieve selected Reporting Dates and Run Keys from the form
        fic_mis_date1 = request.POST.get('fic_mis_date1')
        run_key1 = request.POST.get('run_key1')
        fic_mis_date2 = request.POST.get('fic_mis_date2')
        run_key2 = request.POST.get('run_key2')

        # Validate that both dates and run keys are provided
        if not fic_mis_date1:
            errors['fic_mis_date1'] = 'Please select Reporting Date 1.'
        if not run_key1:
            errors['run_key1'] = 'Please select Run Key 1.'
        if not fic_mis_date2:
            errors['fic_mis_date2'] = 'Please select Reporting Date 2.'
        if not run_key2:
            errors['run_key2'] = 'Please select Run Key 2.'

        # If no errors, proceed with filter
        if not errors:
            request.session['fic_mis_date1'] = fic_mis_date1
            request.session['run_key1'] = run_key1
            request.session['fic_mis_date2'] = fic_mis_date2
            request.session['run_key2'] = run_key2

            # Retrieve filtered data based on the selected main filter values
            ecl_data1 = FCT_Reporting_Lines.objects.filter(fic_mis_date=fic_mis_date1, n_run_key=run_key1)
            ecl_data2 = FCT_Reporting_Lines.objects.filter(fic_mis_date=fic_mis_date2, n_run_key=run_key2)

            # Convert the filtered data into a DataFrame
            ecl_data1_df = pd.DataFrame(list(ecl_data1.values()))
            ecl_data2_df = pd.DataFrame(list(ecl_data2.values()))

            # Ensure both DataFrames use the same set of account numbers
            common_account_numbers = set(ecl_data1_df['n_account_number']).intersection(ecl_data2_df['n_account_number'])
            ecl_data1_df = ecl_data1_df[ecl_data1_df['n_account_number'].isin(common_account_numbers)]
            ecl_data2_df = ecl_data2_df[ecl_data2_df['n_account_number'].isin(common_account_numbers)]

            # Convert date fields to strings for both DataFrames
            for df in [ecl_data1_df, ecl_data2_df]:
                if 'fic_mis_date' in df.columns:
                    df['fic_mis_date'] = df['fic_mis_date'].astype(str)
                if 'd_maturity_date' in df.columns:
                    df['d_maturity_date'] = df['d_maturity_date'].astype(str)

            # Create the directory if it doesn't exist
            if not os.path.exists(CSV_DIR):
                os.makedirs(CSV_DIR)

            # Save the data as CSV files in the session
            csv_filename1 = os.path.join(CSV_DIR, f"ecl_data_{fic_mis_date1}_{run_key1}.csv")
            csv_filename2 = os.path.join(CSV_DIR, f"ecl_data_{fic_mis_date2}_{run_key2}.csv")
            ecl_data1_df.to_csv(csv_filename1, index=False)
            ecl_data2_df.to_csv(csv_filename2, index=False)
            request.session['csv_filename1'] = csv_filename1
            request.session['csv_filename2'] = csv_filename2
            request.session.modified = True

            # Redirect to the sub-filter view
            return redirect('ecl_water_fall_sub_filter')

    # Handle GET request to load Reporting Dates
    fic_mis_dates = FCT_Reporting_Lines.objects.order_by('-fic_mis_date').values_list('fic_mis_date', flat=True).distinct()

    # AJAX request for dynamically updating the Run Key dropdowns
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        field_name = request.GET.get('field')
        field_value = request.GET.get('value')

        if field_name == 'fic_mis_date':
            # Fetch run keys corresponding to the selected FIC MIS date
            n_run_keys = FCT_Reporting_Lines.objects.filter(fic_mis_date=field_value).order_by('-n_run_key').values_list('n_run_key', flat=True).distinct()
            return JsonResponse({'n_run_keys': list(n_run_keys)})

    # If there are errors or it's a GET request, render the form with any errors
    return render(request, 'reports/waterfall_main_filter.html', {
        'fic_mis_dates': fic_mis_dates,
        'errors': errors  # Pass the errors to the template
    })


@login_required
@require_http_methods(["GET", "POST"])
def water_fall_sub_filter_view(request):
    # Retrieve CSV filenames from the session
    csv_filename_1 = request.session.get('csv_filename1')
    csv_filename_2 = request.session.get('csv_filename2')

    # Load CSV files into DataFrames
    ecl_data_df_1 = pd.read_csv(csv_filename_1)
    ecl_data_df_2 = pd.read_csv(csv_filename_2)

    # Merge data based on account number using an inner join
    merged_data = pd.merge(
        ecl_data_df_1,
        ecl_data_df_2,
        on=['n_account_number'],
        how='inner',
        suffixes=('_prev', '_curr')
    )

    # Filter based on account number from the request
    account_number = request.GET.get('n_account_number')
    if account_number:
        merged_data = merged_data[merged_data['n_account_number'] == account_number]

    # Prepare data for the waterfall report
    account_details = {}
    waterfall_data_12m = []
    waterfall_data_lifetime = []
    if not merged_data.empty:
        # Retrieve ECL values and other parameters for both periods
        twelve_month_ecl_prev = merged_data['n_12m_ecl_rcy_prev'].iloc[0]
        twelve_month_ecl_curr = merged_data['n_12m_ecl_rcy_curr'].iloc[0]
        lifetime_ecl_prev = merged_data['n_lifetime_ecl_rcy_prev'].iloc[0]
        lifetime_ecl_curr = merged_data['n_lifetime_ecl_rcy_curr'].iloc[0]

        # Calculate changes for the 12-month ECL waterfall
        stage_change_12m = twelve_month_ecl_curr - twelve_month_ecl_prev if merged_data['n_stage_descr_prev'].iloc[0] != merged_data['n_stage_descr_curr'].iloc[0] else 0
        exposure_change_12m = merged_data['n_exposure_at_default_rcy_curr'].iloc[0] - merged_data['n_exposure_at_default_rcy_prev'].iloc[0]
        pd_change_12m = merged_data['n_twelve_months_pd_curr'].iloc[0] - merged_data['n_twelve_months_pd_prev'].iloc[0]
        lgd_change_12m = merged_data['n_lgd_percent_curr'].iloc[0] - merged_data['n_lgd_percent_prev'].iloc[0]

        # Populate waterfall data for 12-month ECL
        waterfall_data_12m = [
            {"description": "Beginning 12-Month ECL", "impact": twelve_month_ecl_prev},
            {"description": "Change due to Stage (12-Month)", "impact": stage_change_12m},
            {"description": "Change in Exposure at Default (12-Month)", "impact": exposure_change_12m},
            {"description": "Change due to PD (12-Month)", "impact": pd_change_12m},
            {"description": "Change due to LGD (12-Month)", "impact": lgd_change_12m},
            {"description": "Ending 12-Month ECL", "impact": twelve_month_ecl_curr},
        ]

        # Calculate changes for the lifetime ECL waterfall
        stage_change_lifetime = lifetime_ecl_curr - lifetime_ecl_prev if merged_data['n_stage_descr_prev'].iloc[0] != merged_data['n_stage_descr_curr'].iloc[0] else 0
        exposure_change_lifetime = merged_data['n_exposure_at_default_rcy_curr'].iloc[0] - merged_data['n_exposure_at_default_rcy_prev'].iloc[0]
        pd_change_lifetime = merged_data['n_lifetime_pd_curr'].iloc[0] - merged_data['n_lifetime_pd_prev'].iloc[0]
        lgd_change_lifetime = merged_data['n_lgd_percent_curr'].iloc[0] - merged_data['n_lgd_percent_prev'].iloc[0]

        # Populate waterfall data for lifetime ECL
        waterfall_data_lifetime = [
            {"description": "Beginning Lifetime ECL", "impact": lifetime_ecl_prev},
            {"description": "Change due to Stage (Lifetime)", "impact": stage_change_lifetime},
            {"description": "Change in Exposure at Default (Lifetime)", "impact": exposure_change_lifetime},
            {"description": "Change due to PD (Lifetime)", "impact": pd_change_lifetime},
            {"description": "Change due to LGD (Lifetime)", "impact": lgd_change_lifetime},
            {"description": "Ending Lifetime ECL", "impact": lifetime_ecl_curr},
        ]

        # Account details for the selected account
        account_details = {
            'date_prev': merged_data['fic_mis_date_prev'].iloc[0],
            'date_curr': merged_data['fic_mis_date_curr'].iloc[0],
            'n_account_number': merged_data['n_account_number'].iloc[0],
            'exposure_at_default_prev': merged_data['n_exposure_at_default_rcy_prev'].iloc[0],
            'exposure_at_default_curr': merged_data['n_exposure_at_default_rcy_curr'].iloc[0],
            'ifrs_stage_prev': merged_data['n_stage_descr_prev'].iloc[0],
            'ifrs_stage_curr': merged_data['n_stage_descr_curr'].iloc[0],
            'twelve_month_pd_prev': merged_data['n_twelve_months_pd_prev'].iloc[0],
            'twelve_month_pd_curr': merged_data['n_twelve_months_pd_curr'].iloc[0],
            'lifetime_pd_prev': merged_data['n_lifetime_pd_prev'].iloc[0],
            'lifetime_pd_curr': merged_data['n_lifetime_pd_curr'].iloc[0],
            'lgd_prev': merged_data['n_lgd_percent_prev'].iloc[0],
            'lgd_curr': merged_data['n_lgd_percent_curr'].iloc[0],
            'twelve_month_ecl_prev': twelve_month_ecl_prev,
            'twelve_month_ecl_curr': twelve_month_ecl_curr,
            'lifetime_ecl_prev': lifetime_ecl_prev,
            'lifetime_ecl_curr': lifetime_ecl_curr,
            'prod_segment_prev': merged_data['n_prod_segment_prev'].iloc[0],
            'prod_segment_curr': merged_data['n_prod_segment_curr'].iloc[0],
        }

    # Get account options for the dropdown
    account_options = merged_data['n_account_number'].unique().tolist()

    return render(request, 'reports/waterfall_report.html', {
        'account_details': account_details,
        'account_options': account_options,
        'selected_account': account_number,
        'waterfall_data_12m': waterfall_data_12m,
        'waterfall_data_lifetime': waterfall_data_lifetime,
    })



@require_http_methods(["GET", "POST"])
def ecl_graphs_main_filter_view(request):
    # Handle POST request (applying the filter)
    if request.method == 'POST':
        # Retrieve selected Reporting Date and Run Key from the form
        fic_mis_date = request.POST.get('fic_mis_date')
        n_run_key = request.POST.get('n_run_key')
        
        if fic_mis_date and n_run_key:
            # Store the selected filter values in session for later use
            request.session['fic_mis_date'] = fic_mis_date
            request.session['n_run_key'] = n_run_key

            # Retrieve filtered data based on the selected main filter values
            ecl_data = FCT_Reporting_Lines.objects.filter(fic_mis_date=fic_mis_date, n_run_key=n_run_key)

            # Convert the filtered data into a DataFrame
            ecl_data_df = pd.DataFrame(list(ecl_data.values()))

            # Convert date fields to strings
            if 'fic_mis_date' in ecl_data_df.columns:
                ecl_data_df['fic_mis_date'] = ecl_data_df['fic_mis_date'].astype(str)
            if 'd_maturity_date' in ecl_data_df.columns:
                ecl_data_df['d_maturity_date'] = ecl_data_df['d_maturity_date'].astype(str)

            # Create the directory if it doesn't exist
            if not os.path.exists(CSV_DIR):
                os.makedirs(CSV_DIR)

            # Save the data as a CSV file in the session (store the filename)
            csv_filename = os.path.join(CSV_DIR, f"ecl_data_{fic_mis_date}_{n_run_key}.csv")
            ecl_data_df.to_csv(csv_filename, index=False)
            request.session['csv_filename'] = csv_filename

            # Redirect to the sub-filter view
            return redirect('ecl_graphs_sub_filter_view')

    # Handle GET request
    fic_mis_dates = FCT_Reporting_Lines.objects.order_by('-fic_mis_date').values_list('fic_mis_date', flat=True).distinct()

    # Check for AJAX requests to dynamically update run key dropdown
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        field_name = request.GET.get('field')
        field_value = request.GET.get('value')

        if field_name == 'fic_mis_date':
            # Fetch run keys corresponding to the selected FIC MIS date
            n_run_keys = FCT_Reporting_Lines.objects.filter(fic_mis_date=field_value).order_by('-n_run_key').values_list('n_run_key', flat=True).distinct()
            return JsonResponse({'n_run_keys': list(n_run_keys)})

    # Render the main filter template with Reporting Dates
    return render(request, 'reports/ecl_graphs_report_main.html', {'fic_mis_dates': fic_mis_dates})




@login_required
@require_http_methods(["GET", "POST"])
def ecl_graphs_sub_filter_view(request):
    # Retrieve the CSV filename from the session
    csv_filename = request.session.get('csv_filename')

    # If no data is available, redirect to the main filter page
    if not csv_filename or not os.path.exists(csv_filename):
        return redirect('ecl_main_filter_view')

    # Load the data from the CSV file
    ecl_data_df = pd.read_csv(csv_filename)

    # Retrieve sub-filter form fields from the request
    n_prod_segment = request.GET.get('n_prod_segment')
    n_prod_type = request.GET.get('n_prod_type')
    n_stage_descr = request.GET.get('n_stage_descr')
    n_loan_type = request.GET.get('n_loan_type')

    # Apply sub-filters if provided
    if n_prod_segment:
        ecl_data_df = ecl_data_df[ecl_data_df['n_prod_segment'] == n_prod_segment]
    if n_prod_type:
        ecl_data_df = ecl_data_df[ecl_data_df['n_prod_type'] == n_prod_type]
    if n_stage_descr:
        ecl_data_df = ecl_data_df[ecl_data_df['n_stage_descr'] == n_stage_descr]
    if n_loan_type:
        ecl_data_df = ecl_data_df[ecl_data_df['n_loan_type'] == n_loan_type]

    # Retrieve the selected group by field from the request (GET or POST)
    group_by_field = request.GET.get('group_by_field', 'n_stage_descr')  # Default group by 'n_stage_descr'

    # Calculate percentages based on `n_12m_ecl_rcy`
    grouped_df = ecl_data_df.groupby(group_by_field).agg({'n_12m_ecl_rcy': 'sum'}).reset_index()
    total_12m_ecl_rcy = grouped_df['n_12m_ecl_rcy'].sum()
    grouped_df['percentage'] = (grouped_df['n_12m_ecl_rcy'] / total_12m_ecl_rcy) * 100

    # Convert grouped data to a list of dictionaries for the template
    chart_data = grouped_df.to_dict(orient='records')

    # Distinct values for sub-filters
    distinct_prod_segments = list(ecl_data_df['n_prod_segment'].unique())
    distinct_prod_types = list(ecl_data_df['n_prod_type'].unique())
    distinct_stage_descrs = list(ecl_data_df['n_stage_descr'].unique())
    distinct_loan_types = list(ecl_data_df['n_loan_type'].unique())

    # Render the sub-filter view template with calculated percentages
    return render(request, 'reports/ecl_graphs_report_sub.html', {
        'chart_data': chart_data,
        'group_by_field': group_by_field,
        'distinct_prod_segments': distinct_prod_segments,
        'distinct_prod_types': distinct_prod_types,
        'distinct_stage_descrs': distinct_stage_descrs,
        'distinct_loan_types': distinct_loan_types,
    })



@require_http_methods(["GET", "POST"])
def vintage_analysis_main_filter_view(request):
    # Handle POST request (applying the filter)
    if request.method == 'POST':
        # Retrieve selected Reporting Date and Run Key from the form
        fic_mis_date = request.POST.get('fic_mis_date')
        n_run_key = request.POST.get('n_run_key')
        
        if fic_mis_date and n_run_key:
            # Store the selected filter values in session for later use
            request.session['fic_mis_date'] = fic_mis_date
            request.session['n_run_key'] = n_run_key

            # Retrieve filtered data based on the selected main filter values
            ecl_data = FCT_Reporting_Lines.objects.filter(fic_mis_date=fic_mis_date, n_run_key=n_run_key)

            # Convert the filtered data into a DataFrame
            ecl_data_df = pd.DataFrame(list(ecl_data.values()))

            # Convert date fields to strings
            if 'fic_mis_date' in ecl_data_df.columns:
                ecl_data_df['fic_mis_date'] = ecl_data_df['fic_mis_date'].astype(str)
            if 'd_maturity_date' in ecl_data_df.columns:
                ecl_data_df['d_maturity_date'] = ecl_data_df['d_maturity_date'].astype(str)

            # Create the directory if it doesn't exist
            if not os.path.exists(CSV_DIR):
                os.makedirs(CSV_DIR)

            # Save the data as a CSV file in the session (store the filename)
            csv_filename = os.path.join(CSV_DIR, f"ecl_data_{fic_mis_date}_{n_run_key}.csv")
            ecl_data_df.to_csv(csv_filename, index=False)
            request.session['csv_filename'] = csv_filename

            # Redirect to the sub-filter view
            return redirect('vintage_analysis_sub_filter_view')

    # Handle GET request
    fic_mis_dates = FCT_Reporting_Lines.objects.order_by('-fic_mis_date').values_list('fic_mis_date', flat=True).distinct()

    # Check for AJAX requests to dynamically update run key dropdown
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        field_name = request.GET.get('field')
        field_value = request.GET.get('value')

        if field_name == 'fic_mis_date':
            # Fetch run keys corresponding to the selected FIC MIS date
            n_run_keys = FCT_Reporting_Lines.objects.filter(fic_mis_date=field_value).order_by('-n_run_key').values_list('n_run_key', flat=True).distinct()
            return JsonResponse({'n_run_keys': list(n_run_keys)})

    # Render the main filter template with Reporting Dates
    return render(request, 'reports/vintage_analysis_report_main.html', {'fic_mis_dates': fic_mis_dates})




@login_required
@require_http_methods(["GET", "POST"])
def vintage_analysis_sub_filter_view(request):
    # Retrieve the CSV filename from the session
    csv_filename = request.session.get('csv_filename')

    # If no data is available, redirect to the main filter page
    if not csv_filename or not os.path.exists(csv_filename):
        return redirect('ecl_main_filter_view')

    # Load the data from the CSV file
    ecl_data_df = pd.read_csv(csv_filename)

    # Retrieve date range filters from the request
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Apply date range filter if provided
    if start_date:
        ecl_data_df = ecl_data_df[ecl_data_df['d_acct_start_date'] >= start_date]
    if end_date:
        ecl_data_df = ecl_data_df[ecl_data_df['d_acct_start_date'] <= end_date]

    # Retrieve the selected group by field from the request (GET or POST)
    group_by_field = request.GET.get('group_by_field', 'n_stage_descr')  # Default group by 'n_stage_descr'

    # Calculate aggregated data
    grouped_df = ecl_data_df.groupby(group_by_field).agg({
        'n_12m_ecl_rcy': 'sum',  # Total ECL
        'n_account_number': 'count'  # Count of accounts (use the actual column for account IDs)
    }).reset_index()

    # Calculate totals and percentages
    total_12m_ecl_rcy = grouped_df['n_12m_ecl_rcy'].sum()
    total_accounts = grouped_df['n_account_number'].sum()
    grouped_df['percentage'] = (grouped_df['n_12m_ecl_rcy'] / total_12m_ecl_rcy) * 100
    grouped_df['percentage_accounts'] = (grouped_df['n_account_number'] / total_accounts) * 100

    # Add a Grand Total Row
    grand_total = {
        group_by_field: 'Grand Total',
        'n_12m_ecl_rcy': total_12m_ecl_rcy,
        'n_account_number': total_accounts,
        'percentage': 100.0,
        'percentage_accounts': 100.0
    }

    grouped_df = pd.concat([grouped_df, pd.DataFrame([grand_total])], ignore_index=True)

    # Convert grouped data to a list of dictionaries for the template
    # Prepare data for the chart (excluding the Grand Total row)
    chart_data = grouped_df[grouped_df[group_by_field] != 'Grand Total'].to_dict(orient='records')
    table_data = grouped_df.to_dict(orient='records')


    # Distinct values for sub-filters
    distinct_prod_segments = list(ecl_data_df['n_prod_segment'].unique())
    distinct_prod_types = list(ecl_data_df['n_prod_type'].unique())
    distinct_stage_descrs = list(ecl_data_df['n_stage_descr'].unique())
    distinct_loan_types = list(ecl_data_df['n_loan_type'].unique())

    # Convert start_date and end_date to datetime.date objects if they exist
    if start_date:
        try:
            start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
        except ValueError:
            start_date = None  # Handle invalid dates gracefully
    if end_date:
        try:
            end_date = datetime.strptime(end_date, "%Y-%m-%d").date()
        except ValueError:
            end_date = None  # Handle invalid dates gracefully

    print(start_date)
    # Render the sub-filter view template with calculated percentages
    return render(request, 'reports/vintage_analysis_report_sub.html', {
        'chart_data': chart_data,
        'table_data':table_data,
        'group_by_field': group_by_field,
        'distinct_prod_segments': distinct_prod_segments,
        'distinct_prod_types': distinct_prod_types,
        'distinct_stage_descrs': distinct_stage_descrs,
        'distinct_loan_types': distinct_loan_types,
        'start_date': start_date,
        'end_date': end_date,
    })



@login_required
@require_http_methods(["GET", "POST"])
def stage_migration_matrix_main_filter_view(request):
    # Initialize an empty errors dictionary
    errors = {}

    # Handle POST request (applying the filter)
    if request.method == 'POST':
        # Retrieve selected Reporting Dates and Run Keys from the form
        fic_mis_date1 = request.POST.get('fic_mis_date1')
        run_key1 = request.POST.get('run_key1')
        fic_mis_date2 = request.POST.get('fic_mis_date2')
        run_key2 = request.POST.get('run_key2')

        # Validate that both dates and run keys are provided
        if not fic_mis_date1:
            errors['fic_mis_date1'] = 'Please select Reporting Date 1.'
        if not run_key1:
            errors['run_key1'] = 'Please select Run Key 1.'
        if not fic_mis_date2:
            errors['fic_mis_date2'] = 'Please select Reporting Date 2.'
        if not run_key2:
            errors['run_key2'] = 'Please select Run Key 2.'

        # If there are no errors, proceed to filtering
        if not errors:
            # Store the selected filter values in session for later use
            request.session['fic_mis_date1'] = fic_mis_date1
            request.session['run_key1'] = run_key1
            request.session['fic_mis_date2'] = fic_mis_date2
            request.session['run_key2'] = run_key2

            # Retrieve filtered data based on the selected main filter values
            ecl_data1 = FCT_Reporting_Lines.objects.filter(fic_mis_date=fic_mis_date1, n_run_key=run_key1)
            ecl_data2 = FCT_Reporting_Lines.objects.filter(fic_mis_date=fic_mis_date2, n_run_key=run_key2)

            # Convert the filtered data into a DataFrame
            ecl_data1_df = pd.DataFrame(list(ecl_data1.values()))
            ecl_data2_df = pd.DataFrame(list(ecl_data2.values()))

            # Convert date fields to strings for both DataFrames
            for df in [ecl_data1_df, ecl_data2_df]:
                if 'fic_mis_date' in df.columns:
                    df['fic_mis_date'] = df['fic_mis_date'].astype(str)
                if 'd_maturity_date' in df.columns:
                    df['d_maturity_date'] = df['d_maturity_date'].astype(str)

            # Create the directory if it doesn't exist
            if not os.path.exists(CSV_DIR):
                os.makedirs(CSV_DIR)

            # Save the data as CSV files in the session (store the filenames)
            csv_filename1 = os.path.join(CSV_DIR, f"ecl_data_{fic_mis_date1}_{run_key1}.csv")
            csv_filename2 = os.path.join(CSV_DIR, f"ecl_data_{fic_mis_date2}_{run_key2}.csv")
            
            ecl_data1_df.to_csv(csv_filename1, index=False)
            ecl_data2_df.to_csv(csv_filename2, index=False)
            # Store the filenames in the session and explicitly mark the session as modified
            request.session['csv_filename1'] = csv_filename1
            request.session['csv_filename2'] = csv_filename2
            request.session.modified = True
            # Redirect to the sub-filter view
            return redirect('stage_migration_matrix_sub_filter_view')

    # Handle GET request to load Reporting Dates
    fic_mis_dates = FCT_Reporting_Lines.objects.order_by('-fic_mis_date').values_list('fic_mis_date', flat=True).distinct()

    # Check for AJAX requests to dynamically update run key dropdown
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        field_name = request.GET.get('field')
        field_value = request.GET.get('value')

        if field_name == 'fic_mis_date':
            # Fetch run keys corresponding to the selected FIC MIS date
            n_run_keys = FCT_Reporting_Lines.objects.filter(fic_mis_date=field_value).order_by('-n_run_key').values_list('n_run_key', flat=True).distinct()
            return JsonResponse({'n_run_keys': list(n_run_keys)})

    # If there are errors or it's a GET request, render the form with any errors
    return render(request, 'reports/stage_migration_matrix_main_filter.html', {
        'fic_mis_dates': fic_mis_dates,
        'errors': errors  # Pass the errors to the template
    })



@login_required
@require_http_methods(["GET", "POST"])
def stage_migration_matrix_sub_filter_view(request):
    # Retrieve CSV filenames from the session
    csv_filename_1 = request.session.get('csv_filename1')
    csv_filename_2 = request.session.get('csv_filename2')

    # Validate file existence
    if not csv_filename_1 or not os.path.exists(csv_filename_1):
        messages.error(request, "Dataset 1 is missing or invalid.")
        return redirect('ecl_main_filter_view')

    if not csv_filename_2 or not os.path.exists(csv_filename_2):
        messages.error(request, "Dataset 2 is missing or invalid.")
        return redirect('ecl_main_filter_view')

    # Load datasets
    ecl_data_1 = pd.read_csv(csv_filename_1)
    ecl_data_2 = pd.read_csv(csv_filename_2)

    # Combine datasets for distinct filters
    ecl_data_combined = pd.concat([ecl_data_1, ecl_data_2])

    # Distinct values for filters
    distinct_prod_segments = list(ecl_data_combined['n_prod_segment'].dropna().unique())
    distinct_prod_types = list(ecl_data_combined['n_prod_type'].dropna().unique())
    distinct_stage_descrs = list(ecl_data_combined['n_stage_descr'].dropna().unique())
    distinct_loan_types = list(ecl_data_combined['n_loan_type'].dropna().unique())
    distinct_customer_types = list(ecl_data_combined['n_party_type'].dropna().unique())

    # Extract the actual dates from the datasets
    date_1 = ecl_data_1['fic_mis_date'].iloc[0] if 'fic_mis_date' in ecl_data_1.columns else "Unknown"
    date_2 = ecl_data_2['fic_mis_date'].iloc[0] if 'fic_mis_date' in ecl_data_2.columns else "Unknown"

    # Apply filters based on user input
    v_ccy_code = request.GET.get('v_ccy_code')
    n_prod_segment = request.GET.get('n_prod_segment')
    n_prod_type = request.GET.get('n_prod_type')
    n_loan_type = request.GET.get('n_loan_type')
    customer_type = request.GET.get('customer_type')

    filters = {
        'v_ccy_code': v_ccy_code,
        'n_prod_segment': n_prod_segment,
        'n_prod_type': n_prod_type,
        'n_loan_type': n_loan_type,
        'customer_type': customer_type,
    }

    for column, value in filters.items():
        if value:
            ecl_data_1 = ecl_data_1[ecl_data_1[column] == value]
            ecl_data_2 = ecl_data_2[ecl_data_2[column] == value]

    # Merge datasets on `n_account_number` for stage comparison
    merged_data = pd.merge(
        ecl_data_1[['n_account_number', 'n_stage_descr']],
        ecl_data_2[['n_account_number', 'n_stage_descr']],
        on='n_account_number',
        how='inner',
        suffixes=('_start', '_end')
    )

    # Create Stage Migration Matrix
    migration_matrix_counts = pd.crosstab(
        merged_data['n_stage_descr_start'],
        merged_data['n_stage_descr_end'],
        margins=True,
        margins_name="Total"
    )

    # Calculate percentages
    migration_matrix_percentages = migration_matrix_counts.div(
        migration_matrix_counts.loc["Total"], axis=1
    ) * 100

    # Prepare data for the matrix table
    stages = ["Stage 1", "Stage 2", "Stage 3"]
    migration_data = []

    for from_stage in migration_matrix_counts.index[:-1]:  # Exclude 'Total'
        row_data = {"from_stage": from_stage}
        for to_stage in stages:
            count = migration_matrix_counts.loc[from_stage, to_stage] if to_stage in migration_matrix_counts.columns else 0
            percentage = migration_matrix_percentages.loc[from_stage, to_stage] if to_stage in migration_matrix_percentages.columns else 0
            row_data[to_stage] = {"count": int(count), "percentage": percentage}
        migration_data.append(row_data)

    # Calculate the number of accounts and percentages for each date
    date_1_data = ecl_data_1['n_stage_descr'].value_counts(normalize=False).to_dict()
    date_2_data = ecl_data_2['n_stage_descr'].value_counts(normalize=False).to_dict()

    date_1_percentages = ecl_data_1['n_stage_descr'].value_counts(normalize=True) * 100
    date_2_percentages = ecl_data_2['n_stage_descr'].value_counts(normalize=True) * 100

    # Calculate totals for each date
    total_date_1_count = sum(date_1_data.values())
    total_date_2_count = sum(date_2_data.values())
    total_date_1_percentage = 100.0
    total_date_2_percentage = 100.0

    account_summary = {
        "date_1": [{"stage": stage, "count": date_1_data.get(stage, 0), "percentage": date_1_percentages.get(stage, 0)} for stage in stages],
        "date_2": [{"stage": stage, "count": date_2_data.get(stage, 0), "percentage": date_2_percentages.get(stage, 0)} for stage in stages],
        "total_date_1": {"count": total_date_1_count, "percentage": total_date_1_percentage},
        "total_date_2": {"count": total_date_2_count, "percentage": total_date_2_percentage},
    }

    # Identify new accounts in date_2 but not in date_1
    new_accounts = ecl_data_2[~ecl_data_2['n_account_number'].isin(ecl_data_1['n_account_number'])]
    new_accounts_summary = new_accounts['n_stage_descr'].value_counts().reindex(stages, fill_value=0).reset_index()
    new_accounts_summary.columns = ['stage', 'count']

    if date_1:
        try:
            date_1 = datetime.strptime(date_1, "%Y-%m-%d").date()
        except ValueError:
            date_1 = None  # Handle invalid dates gracefully
    if date_2:
        try:
            date_2 = datetime.strptime(date_2, "%Y-%m-%d").date()
        except ValueError:
            date_2 = None  # Handle invalid dates gracefully

    # Render the template with data
    return render(request, 'reports/stage_migration_matrix_report_sub.html', {
        'stages': stages,
        'migration_data': migration_data,
        'account_summary': account_summary,
        'new_accounts_summary': new_accounts_summary.to_dict(orient='records'),
        'date_1': date_1,
        'date_2': date_2,
        'filters': filters,
        'distinct_prod_segments': distinct_prod_segments,
        'distinct_prod_types': distinct_prod_types,
        'distinct_stage_descrs': distinct_stage_descrs,
        'distinct_loan_types': distinct_loan_types,
        'distinct_customer_types': distinct_customer_types,
    })
